"""Exam results and Excel report generation (2-sheet mẫu trường, AD-68)."""

from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from datetime import date as date_type

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Answer, Candidate, Exam, ExamSession
from app.models.enums import SessionStatus
from app.models.room import Room
from app.services import session_service

_SUBMITTED = (SessionStatus.SUBMITTED.value, SessionStatus.TIMEOUT.value)

_TRAILING_NUM = re.compile(r"(\d+)\s*$")


def _code_number(code: str) -> int | None:
    """Số ở CUỐI mã câu hỏi ("item_Q-153" → 153, "111101" → 111101); None nếu
    mã không kết thúc bằng số — dùng sắp cột báo cáo theo cấu trúc đề Q-1..Q-n."""
    m = _TRAILING_NUM.search(code or "")
    return int(m.group(1)) if m else None


async def get_answer_key(redis, sitting) -> dict[str, dict]:
    """{question_id: {text, correct_option}} in the original question order, for a
    sitting (buổi thi). Resolution order (AD-47):
      1. sitting.report_snapshot (permanent — written at QTI import, survives the
         end-of-sitting payload purge).
      2. Redis :report cache.
      3. Live Redis payload (sitting still open, never snapshotted before).
    """
    if sitting is not None and sitting.report_snapshot:
        return {
            q["id"]: {"code": q.get("code", ""), "text": q.get("text", ""),
                      "correct_option": q.get("correct_option")}
            for q in sitting.report_snapshot
        }
    sid = sitting.id if sitting else None
    raw = await redis.get(session_service.report_cache_key(sid))
    if raw:
        return json.loads(raw)
    payload = await session_service.get_sitting_payload(redis, sid)
    if not payload:
        return {}
    key = {q["id"]: {"code": q.get("code", ""), "text": q["text"],
                     "correct_option": q["correct_option"]}
           for q in payload["questions"]}
    await redis.set(session_service.report_cache_key(sid), json.dumps(key, ensure_ascii=False))
    return key


async def build_report(db: AsyncSession, sitting, exam_name: str = "") -> dict:
    """Build the results report for a single sitting (buổi thi, AD-47/AD-68).

    Trả về:
      - meta: thông tin buổi thi
      - rows: MỌI thí sinh đăng ký kỳ thi, có hoặc không có phiên
      - questions: danh sách câu hỏi (thứ tự gốc) kèm đáp án đúng
    """
    from app.core.redis import redis_client
    answer_key = await get_answer_key(redis_client, sitting)
    # Cột câu hỏi sắp theo SỐ ở cuối mã câu (Q-1..Q-n) — nhà cung cấp hay xuất
    # file test XML với item-ref xáo trộn nên thứ tự snapshot không phải "cấu
    # trúc đề" người đọc muốn tra (13-07). Chỉ sắp khi MỌI câu đều có mã kết
    # thúc bằng số; đề không mã giữ nguyên thứ tự gốc. Đáp án map theo
    # question_id nên đổi thứ tự cột không ảnh hưởng điểm/đối chiếu.
    items = list(answer_key.items())
    order_keys = [_code_number(v.get("code", "")) for _, v in items]
    if items and all(k is not None for k in order_keys):
        items.sort(key=lambda kv: _code_number(kv[1].get("code", "")))
    qids = [qid for qid, _ in items]
    questions = [
        {"index": i + 1, "code": v.get("code", ""), "correct_option": v["correct_option"]}
        for i, (_, v) in enumerate(items)
    ]

    # Lấy ngày thi từ kỳ thi cha
    exam = await db.get(Exam, sitting.exam_id)
    raw_date = None
    if exam and exam.exam_date:
        raw_date = exam.exam_date
    elif sitting.scheduled_date:
        raw_date = sitting.scheduled_date
    exam_date_str = raw_date.strftime("%d/%m/%Y") if raw_date else ""

    # Tất cả thí sinh đăng ký kỳ thi (không chỉ người có phiên)
    cand_result = await db.execute(
        select(Candidate, Room.name.label("room_name"))
        .outerjoin(Room, Candidate.room_id == Room.id)
        .where(Candidate.exam_id == sitting.exam_id)
    )
    candidates_with_rooms = cand_result.all()

    # Phiên thi của buổi này — map theo candidate_id
    sess_result = await db.execute(
        select(ExamSession).where(ExamSession.sitting_id == sitting.id)
    )
    sessions = sess_result.scalars().all()
    session_by_candidate: dict[str, ExamSession] = {
        str(s.candidate_id): s for s in sessions
    }

    # Đáp án của từng phiên — map [session_id][question_id] = option letter
    answers_by_session: dict[str, dict[str, str]] = defaultdict(dict)
    all_session_ids = [str(s.id) for s in sessions]
    if all_session_ids:
        ans_result = await db.execute(
            select(Answer.session_id, Answer.question_id, Answer.selected_option)
            .where(Answer.session_id.in_(all_session_ids))
        )
        for sid, qid, opt in ans_result.all():
            answers_by_session[str(sid)][str(qid)] = opt or ""

    rows = []
    for cand, room_name in candidates_with_rooms:
        sess = session_by_candidate.get(str(cand.id))

        # Tách họ đệm / tên
        parts = cand.full_name.rsplit(" ", 1)
        ho_dem = parts[0] if len(parts) == 2 else ""
        ten = parts[-1]

        # Trạng thái và điểm
        if sess is None:
            status = "absent"
            score = None
            total = None
        elif sess.status in _SUBMITTED:
            status = sess.status
            score = float(sess.score) if sess.score is not None else None
            total = sess.total_correct
        else:
            status = sess.status
            score = None
            total = None

        # Đáp án đã chọn theo thứ tự câu gốc (letter | "")
        sid = str(sess.id) if sess else None
        answers = [
            answers_by_session.get(sid, {}).get(qid, "") if sid else ""
            for qid in qids
        ]

        rows.append({
            "cccd": cand.cccd,
            "ho_dem": ho_dem,
            "ten": ten,
            "room_name": room_name or "",
            "birth_date": cand.birth_date,
            "score": score,
            "total_correct": total,
            "status": status,
            "answers": answers,
            # stt gán sau khi sắp xếp
        })

    # Sắp xếp theo tên rồi họ đệm (khớp mẫu trường)
    rows.sort(key=lambda r: (r["ten"], r["ho_dem"]))
    for i, row in enumerate(rows):
        row["stt"] = i + 1

    return {
        "meta": {
            "exam_name": exam_name or (exam.name if exam else ""),
            "sitting_name": sitting.name,
            "exam_date": exam_date_str,
            "question_count": len(qids),
        },
        "rows": rows,
        "questions": questions,
    }


# --- Excel 2 sheet -----------------------------------------------------------

def _fmt_date(d) -> str:
    """Chuyển date object hoặc ISO string sang dd/mm/yyyy."""
    if d is None:
        return ""
    if isinstance(d, date_type):
        return d.strftime("%d/%m/%Y")
    # fallback: chuỗi ISO yyyy-mm-dd
    try:
        parts = str(d).split("-")
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:  # noqa: BLE001
        return str(d)


def export_excel(report: dict, exam_name: str | None = None) -> bytes:
    """Xuất báo cáo dạng xlsx 2 sheet: 'Kết quả' và 'Đáp án'.

    Đúng mẫu nhà trường (AD-68):
      - 'Kết quả': danh sách thí sinh + điểm + số câu đúng.
      - 'Đáp án': bảng chọn từng câu (theo thứ tự đề gốc) kèm dòng đáp án đúng.
    """
    meta = report["meta"]
    rows = report["rows"]
    questions = report["questions"]
    name = exam_name or meta.get("exam_name", "")
    exam_date = meta.get("exam_date", "")
    n_q = len(questions)

    wb = Workbook()

    # ── Sheet 1: Kết quả ─────────────────────────────────────────────────────
    ws_kq = wb.active
    ws_kq.title = "Kết quả"
    bold = Font(bold=True)

    ws_kq.append(["DANH SÁCH KẾT QUẢ BÀI THI"])
    ws_kq.cell(row=1, column=1).font = bold
    ws_kq.append([f"KẾT QUẢ THI MÔN: {name}"])
    ws_kq.cell(row=2, column=1).font = bold
    ws_kq.append([f"Ngày thi: {exam_date}"])
    ws_kq.append([])  # dòng trống

    headers_kq = ["STT", "Số báo danh", "Họ đệm", "Tên", "Tổ", "Ngày sinh", "Điểm", "Số câu đúng"]
    ws_kq.append(headers_kq)
    for col_idx in range(1, len(headers_kq) + 1):
        ws_kq.cell(row=5, column=col_idx).font = bold

    for row in rows:
        if row["status"] == "absent":
            diem_cell = "Vắng"
        elif row["score"] is not None:
            diem_cell = row["score"]
        else:
            diem_cell = ""
        so_cau = row["total_correct"] if row["total_correct"] is not None else ""
        ws_kq.append([
            row["stt"],
            row["cccd"],
            row["ho_dem"],
            row["ten"],
            row["room_name"],
            _fmt_date(row["birth_date"]),
            diem_cell,
            so_cau,
        ])

    # Độ rộng cột
    for col, width in zip("ABCDEFGH", [6, 16, 22, 12, 14, 14, 10, 14]):
        ws_kq.column_dimensions[col].width = width

    # ── Sheet 2: Đáp án ──────────────────────────────────────────────────────
    ws_da = wb.create_sheet("Đáp án")

    ws_da.append(["DANH SÁCH KẾT QUẢ BÀI THI"])
    ws_da.cell(row=1, column=1).font = bold
    ws_da.append([f"Số câu hỏi: {n_q}"])
    ws_da.cell(row=2, column=1).font = bold
    ws_da.append([f"Tên môn học: {name}"])
    ws_da.cell(row=3, column=1).font = bold
    ws_da.append([])  # dòng trống

    # Header: STT, Mã sinh viên, Họ đệm, Tên, Ngày sinh, 1, 2, ..., N (dòng 5)
    headers_da = ["STT", "Mã sinh viên", "Họ đệm", "Tên", "Ngày sinh"] + list(range(1, n_q + 1))
    ws_da.append(headers_da)
    for col_idx in range(1, len(headers_da) + 1):
        ws_da.cell(row=5, column=col_idx).font = bold

    # Dòng "Mã câu hỏi" (dòng 6): nhãn ở cột Mã sinh viên (B), mỗi cột câu = mã câu
    # hỏi gốc (identifier trong file QTI). Đề nạp trước bản này chưa lưu mã → để trống.
    code_row = ["", "Mã câu hỏi", "", "", ""] + [q.get("code", "") for q in questions]
    ws_da.append(code_row)
    ws_da.cell(row=6, column=2).font = bold

    # Dòng "Đáp án đúng" (dòng 7): nhãn ở cột Mã sinh viên (B), cột câu = đáp án đúng
    correct_row = ["", "Đáp án đúng", "", "", ""] + [q["correct_option"] for q in questions]
    ws_da.append(correct_row)
    ws_da.cell(row=7, column=2).font = bold

    # Dữ liệu từng thí sinh (dòng 8+)
    for row in rows:
        ws_da.append(
            [row["stt"], row["cccd"], row["ho_dem"], row["ten"], _fmt_date(row["birth_date"])]
            + row["answers"]
        )

    # Độ rộng cột cố định
    ws_da.column_dimensions["A"].width = 14
    ws_da.column_dimensions["B"].width = 16
    ws_da.column_dimensions["C"].width = 22
    ws_da.column_dimensions["D"].width = 12
    ws_da.column_dimensions["E"].width = 14
    for i in range(6, 6 + n_q):
        from openpyxl.utils import get_column_letter
        ws_da.column_dimensions[get_column_letter(i)].width = 6

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_excel_encrypted_zip(report: dict, password: str, filename: str = "report.xlsx") -> bytes:
    """Build the .xlsx and wrap it in a password-protected AES-encrypted ZIP.
    Standard 7-Zip / WinRAR / macOS Archive Utility can open it given the
    password — no special viewer needed."""
    import pyzipper
    xlsx_bytes = export_excel(report)
    bio = io.BytesIO()
    with pyzipper.AESZipFile(
        bio, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES,
    ) as zf:
        zf.setpassword(password.encode("utf-8"))
        zf.writestr(filename, xlsx_bytes)
    return bio.getvalue()
