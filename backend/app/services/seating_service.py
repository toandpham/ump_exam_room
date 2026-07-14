"""Seating list (danh sách phòng/ghế) export — Excel + PDF (AD-47).

``rooms`` is a list of:
  {"room_name": str, "proctor_name": str|None,
   "candidates": [{"seat": int, "full_name": str, "cccd": str, "unit": str}, ...]}
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Font Unicode cho tiếng Việt trong PDF
_FONT = "DejaVu"
_FONT_BOLD = "DejaVu-Bold"
try:
    pdfmetrics.registerFont(TTFont(_FONT, "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
    pdfmetrics.registerFont(
        TTFont(_FONT_BOLD, "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
    )
except Exception:  # noqa: BLE001
    _FONT = "Helvetica"
    _FONT_BOLD = "Helvetica-Bold"


def export_excel(exam_name: str, rooms: list[dict]) -> bytes:
    wb = Workbook()
    first = True
    for room in rooms:
        title = (room["room_name"] or "Phòng")[:28]
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
            first = False
        ws.append([f"{exam_name} — {room['room_name']}"])
        ws.append([f"Giám thị: {room.get('proctor_name') or '—'}"])
        ws.append([])
        header_row = ws.max_row + 1
        ws.append(["STT", "Họ tên", "CCCD", "Đơn vị"])
        font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="2F5496")
        for i in range(1, 5):
            cell = ws.cell(row=header_row, column=i)
            cell.font = font
            cell.fill = fill
        for idx, c in enumerate(room["candidates"], start=1):
            ws.append([idx, c["full_name"], c["cccd"], c["unit"]])
            ws.cell(row=ws.max_row, column=3).number_format = "@"
        for i, w in enumerate([8, 30, 16, 30], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    if first:  # no rooms → keep a placeholder sheet
        wb.active.title = "Phòng thi"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_pdf(exam_name: str, rooms: list[dict]) -> bytes:
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, title="Danh sách phòng thi")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontName=_FONT_BOLD)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName=_FONT_BOLD)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontName=_FONT)

    elements = [Paragraph("DANH SÁCH PHÒNG THI", title_style),
                Paragraph(exam_name, normal), Spacer(1, 12)]
    for room in rooms:
        elements.append(Paragraph(
            f"{room['room_name']} — Giám thị: {room.get('proctor_name') or '—'}", h2))
        data = [["STT", "Họ tên", "CCCD", "Đơn vị"]]
        for idx, c in enumerate(room["candidates"], start=1):
            data.append([str(idx), c["full_name"], c["cccd"], c["unit"]])
        t = Table(data, colWidths=[40, 180, 90, 160])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), _FONT),
            ("FONTNAME", (0, 0), (-1, 0), _FONT_BOLD),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elements += [t, Spacer(1, 16)]
    doc.build(elements)
    return bio.getvalue()
