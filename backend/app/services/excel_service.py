"""Excel template, parsing/validation, and export for candidates.

Template columns (row 1 = header, data from row 2):
  A CCCD/Hộ chiếu | B Họ tên | C Ngày sinh | D Đơn vị | E Năm TN | F Ngành |
  G Đối tượng | H Lần dự thi | I Phòng (tuỳ chọn — tên phòng để tự chia phòng)
The login id may be a CCCD (12 digits) or a passport (6–9 alnum); the type is
auto-detected (AD-58). Photos are uploaded separately (single or ZIP by id).
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.identifier import classify_identifier

HEADERS = [
    "CCCD/Hộ chiếu",
    "Họ tên",
    "Ngày sinh (YYYY-MM-DD)",
    "Đơn vị",
    "Năm tốt nghiệp",
    "Ngành",
    "Đối tượng",
    "Lần dự thi",
    "Phòng (tuỳ chọn)",
]
COL_WIDTHS = [16, 24, 22, 28, 14, 20, 16, 12, 16]


def build_template() -> bytes:
    """Generate the candidate import template (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ThiSinh"
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS[idx - 1]
    # Example rows: a CCCD and a passport (foreign candidate).
    ws.append(["079200000099", "Nguyễn Văn Mẫu", "1995-05-20", "Đơn vị A",
               2017, "Y đa khoa", "Đối tượng 1", 1, "Phòng 1"])
    ws.append(["C1234567", "John Smith", "1990-08-15", "Đơn vị B",
               2015, "Y đa khoa", "Đối tượng 2", 1, "Phòng 1"])
    # Force the CCCD/passport column to TEXT so leading zeros + letters are kept.
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "@"
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_rows(content: bytes) -> list[tuple[int, tuple]]:
    """Return non-empty data rows as (row_number, values_tuple)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[tuple[int, tuple]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        rows.append((row_number, values))
    wb.close()
    return rows


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_cccd(value: Any) -> str:
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(12)
    return _cell_str(value)


def _parse_birth_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError


def validate_row(values: tuple) -> tuple[dict, list[str]]:
    """Validate a single data row. Returns (parsed_data, errors).

    parsed_data uses ISO strings for dates so it is JSON-serializable for the
    preview/commit Redis round-trip.
    """
    errors: list[str] = []
    padded = list(values) + [None] * (9 - len(values))
    (cccd_raw, name_raw, dob_raw, unit_raw, grad_raw, major_raw, cat_raw,
     attempt_raw, room_raw) = padded[:9]

    raw_id = _parse_cccd(cccd_raw)
    id_type = "cccd"
    try:
        cccd, id_type = classify_identifier(raw_id)   # CCCD or passport (AD-58)
    except ValueError as exc:
        cccd = raw_id
        errors.append(str(exc))

    full_name = _cell_str(name_raw)
    if not full_name:
        errors.append("Họ tên không được để trống")

    birth_iso: str | None = None
    try:
        birth_iso = _parse_birth_date(dob_raw).isoformat()
    except ValueError:
        errors.append("Ngày sinh không hợp lệ (dùng YYYY-MM-DD hoặc DD/MM/YYYY)")

    unit = _cell_str(unit_raw)
    if not unit:
        errors.append("Đơn vị không được để trống")

    graduation_year: int | None = None
    grad_str = _cell_str(grad_raw)
    if grad_str:
        try:
            graduation_year = int(float(grad_str))
            if not (1900 <= graduation_year <= 2100):
                errors.append("Năm tốt nghiệp không hợp lệ")
        except ValueError:
            errors.append("Năm tốt nghiệp phải là số")

    major = _cell_str(major_raw) or None

    category = _cell_str(cat_raw)
    if not category:
        errors.append("Đối tượng không được để trống")

    attempt_number = 1
    attempt_str = _cell_str(attempt_raw)
    if attempt_str:
        try:
            attempt_number = int(float(attempt_str))
            if attempt_number < 1:
                errors.append("Lần dự thi phải >= 1")
        except ValueError:
            errors.append("Lần dự thi phải là số")

    data = {
        "cccd": cccd,
        "id_type": id_type,
        "full_name": full_name,
        "birth_date": birth_iso,
        "unit": unit,
        "graduation_year": graduation_year,
        "major": major,
        "category": category,
        "attempt_number": attempt_number,
        "room_name": _cell_str(room_raw) or None,
    }
    return data, errors


def export_candidates(rows: list[dict]) -> bytes:
    """Export candidate rows to .xlsx. Each row is a plain dict of fields."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ThiSinh"
    headers = ["CCCD", "Họ tên", "Ngày sinh", "Đơn vị", "Năm TN", "Ngành",
               "Đối tượng", "Lần dự thi", "Phòng", "Kỳ thi", "Có hình"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=idx).font = header_font
        ws.cell(row=1, column=idx).fill = header_fill
        ws.column_dimensions[get_column_letter(idx)].width = 18
    for r in rows:
        ws.append([
            r.get("cccd"), r.get("full_name"), r.get("birth_date"), r.get("unit"),
            r.get("graduation_year"), r.get("major"), r.get("category"),
            r.get("attempt_number"), r.get("room_name"), r.get("exam_name"),
            "Có" if r.get("photo_path") else "Không",
        ])
        ws.cell(row=ws.max_row, column=1).number_format = "@"
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
