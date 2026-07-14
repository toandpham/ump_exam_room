"""Báo cáo 2 sheet đúng mẫu trường (AD-68): Kết quả + Đáp án."""
import io
import pytest
from openpyxl import load_workbook

from app.services import report_service
from tests.conftest import auth, fast_forward_start

pytestmark = pytest.mark.asyncio

# ─── helpers ──────────────────────────────────────────────────────────────────


def _row_values(ws, row_num):
    """Trả list giá trị của dòng row_num (1-indexed) trong worksheet."""
    return [c.value for c in ws[row_num]]


# ─── tests ────────────────────────────────────────────────────────────────────


async def test_export_excel_has_two_sheets_and_absent(client, factory, db):
    """2 sheet đúng tên; dòng header đúng; thí sinh vắng có ô 'Vắng';
    dòng 'Đáp án đúng' ghi đúng đáp án; thí sinh nộp bài có cột đáp án đúng."""

    admin, ptok = await factory.admin()
    exam, sitting, payload = await factory.active_exam(
        [{"text": "Q1", "correct": "A", "code": "111101"},
         {"text": "Q2", "correct": "B", "code": "111102"}],
        owner_id=admin.id,
    )
    q1id = payload["questions"][0]["id"]
    q2id = payload["questions"][1]["id"]  # noqa: F841 — dùng để kiểm tra q2 không chọn

    # Candidate 1: đăng nhập, xác nhận, phát đề, bắt đầu, trả lời Q1, nộp bài
    cand1 = await factory.candidate(exam.id)
    tok1 = (await client.post("/api/exam/auth/login", json={"cccd": cand1.cccd})).json()["token"]
    ch1 = auth(tok1)
    await client.post("/api/exam/auth/confirm", headers=ch1)
    await client.post(f"/api/admin/sittings/{sitting.id}/start", headers=auth(ptok))
    await fast_forward_start(sitting.id)  # SP-2c: bỏ qua đếm ngược để trả lời ngay
    await client.post("/api/exam/answer",
                      json={"question_id": q1id, "selected_option": "A"}, headers=ch1)
    await client.post("/api/exam/submit", headers=ch1)

    # Candidate 2: chỉ đăng ký, không đăng nhập → vắng trong buổi này
    cand2 = await factory.candidate(exam.id)

    # ── build_report (không cần redis tường minh — dùng report_snapshot) ──────
    report = await report_service.build_report(db, sitting, exam.name)

    # Cấu trúc trả về đúng key
    assert "meta" in report and "rows" in report and "questions" in report

    meta = report["meta"]
    assert meta["question_count"] == 2

    # questions ghi đúng đáp án đúng
    assert [q["correct_option"] for q in report["questions"]] == ["A", "B"]

    # rows gồm đủ 2 thí sinh
    assert len(report["rows"]) == 2

    # Tìm row của từng thí sinh
    by_cccd = {r["cccd"]: r for r in report["rows"]}

    r1 = by_cccd[cand1.cccd]
    assert r1["status"] in ("submitted", "timeout")
    assert r1["total_correct"] == 1      # chỉ Q1 đúng
    assert r1["answers"][0] == "A"       # Q1 chọn A
    assert r1["answers"][1] == ""        # Q2 không trả lời

    r2 = by_cccd[cand2.cccd]
    assert r2["status"] == "absent"
    assert r2["score"] is None
    assert r2["total_correct"] is None
    assert r2["answers"] == ["", ""]    # vắng → không có đáp án

    # ── export_excel ──────────────────────────────────────────────────────────
    data = report_service.export_excel(report, exam.name)
    wb = load_workbook(io.BytesIO(data))

    # Đúng 2 sheet theo thứ tự
    assert wb.sheetnames == ["Kết quả", "Đáp án"]

    # ── Sheet "Kết quả" ───────────────────────────────────────────────────────
    ws_kq = wb["Kết quả"]

    # Dòng 1–3: tiêu đề; dòng 4: trống; dòng 5: header
    header_kq = _row_values(ws_kq, 5)
    assert "Số câu đúng" in header_kq
    assert "Điểm" in header_kq
    assert "STT" in header_kq

    # Tìm index cột CCCD ("Số báo danh") và "Điểm"
    col_cccd = header_kq.index("Số báo danh") + 1   # 1-indexed
    col_diem = header_kq.index("Điểm") + 1

    # Dòng dữ liệu bắt đầu từ row 6
    diem_by_cccd: dict[str, object] = {}
    for row in ws_kq.iter_rows(min_row=6, values_only=True):
        if row[col_cccd - 1] is not None:
            diem_by_cccd[row[col_cccd - 1]] = row[col_diem - 1]

    assert diem_by_cccd.get(cand2.cccd) == "Vắng"
    # Thí sinh đã nộp: ô điểm là số (không phải chuỗi "Vắng")
    assert diem_by_cccd.get(cand1.cccd) != "Vắng"

    # ── Sheet "Đáp án" ────────────────────────────────────────────────────────
    ws_da = wb["Đáp án"]

    # Dòng 5: header; dòng 6: "Mã câu hỏi"; dòng 7: "Đáp án đúng"; dòng 8+: thí sinh
    header_da = _row_values(ws_da, 5)
    assert "Mã sinh viên" in header_da
    # Cột câu hỏi bắt đầu từ index 5 (0-based), tức 1 và 2 trong header
    assert header_da[5] == 1
    assert header_da[6] == 2

    # Dòng 6 = "Mã câu hỏi" (nhãn ở cột B = Mã sinh viên), mỗi câu = mã QTI gốc
    code_row = _row_values(ws_da, 6)
    assert code_row[1] == "Mã câu hỏi"
    assert code_row[5] == "111101"   # Q1 mã câu hỏi
    assert code_row[6] == "111102"   # Q2 mã câu hỏi

    # Dòng 7 = "Đáp án đúng" (nhãn ở cột B)
    correct_row = _row_values(ws_da, 7)
    assert correct_row[1] == "Đáp án đúng"
    assert correct_row[5] == "A"     # Q1 đáp án đúng
    assert correct_row[6] == "B"     # Q2 đáp án đúng

    # Tìm dòng của cand1 trong sheet Đáp án (dòng 8+)
    col_cccd_da = header_da.index("Mã sinh viên")  # 0-based
    for row in ws_da.iter_rows(min_row=8, values_only=True):
        if row[col_cccd_da] == cand1.cccd:
            assert row[5] == "A"               # Q1 chọn A
            assert row[6] in ("", None)     # Q2 không chọn (openpyxl đọc "" → None)
            break
    else:
        pytest.fail(f"Không tìm thấy thí sinh {cand1.cccd} trong sheet Đáp án")


async def test_build_report_sorts_by_ten_ho_dem(factory, db):
    """Rows được sắp theo (tên, họ đệm) rồi mới gán STT."""
    _, _ = await factory.admin()
    exam, sitting, _ = await factory.active_exam(
        [{"text": "Q1", "correct": "A"}],
    )
    await factory.candidate(exam.id, cccd="900000000001")
    await factory.candidate(exam.id, cccd="900000000002")
    # Cả 2 đều vắng (không có phiên)
    report = await report_service.build_report(db, sitting, exam.name)
    stts = [r["stt"] for r in report["rows"]]
    assert stts == sorted(stts)  # STT liên tục tăng dần


async def test_export_excel_no_questions(factory, db):
    """Xuất được bảng khi buổi chưa có đề (0 câu)."""
    from app.models.enums import SittingStatus
    import uuid as _uuid
    from app.models import Exam as _Exam
    from app.models.sitting import Sitting as _Sitting
    from datetime import date
    from app.models.enums import ExamStatus
    from sqlalchemy.ext.asyncio import AsyncSession

    # Dùng factory để tạo exam, sitting trống
    admin, _ = await factory.admin()
    exam, sitting, _ = await factory.active_exam([], owner_id=admin.id)

    report = await report_service.build_report(db, sitting, exam.name)
    assert report["meta"]["question_count"] == 0
    data = report_service.export_excel(report, exam.name)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Kết quả", "Đáp án"]


async def test_report_orders_questions_by_code_number(client, factory, db):
    """Cột câu hỏi sắp theo SỐ ở cuối mã câu (Q-1..Q-n) thay vì thứ tự item-ref
    trong file test XML — nhà cung cấp hay xuất test xml xáo trộn (13-07:
    item_Q-153, item_Q-81… trong khi cấu trúc đề là Q-1→Q-n). Đáp án thí sinh
    phải dịch cột theo đúng thứ tự mới."""
    admin, ptok = await factory.admin()
    exam, sitting, payload = await factory.active_exam(
        [{"text": "QA", "correct": "A", "code": "item_Q-153"},
         {"text": "QB", "correct": "B", "code": "item_Q-2"},
         {"text": "QC", "correct": "C", "code": "item_Q-81"}],
        owner_id=admin.id,
    )
    q153 = payload["questions"][0]["id"]

    cand = await factory.candidate(exam.id)
    tok = (await client.post("/api/exam/auth/login", json={"cccd": cand.cccd})).json()["token"]
    ch = auth(tok)
    await client.post("/api/exam/auth/confirm", headers=ch)
    await client.post(f"/api/admin/sittings/{sitting.id}/start", headers=auth(ptok))
    await fast_forward_start(sitting.id)
    await client.post("/api/exam/answer",
                      json={"question_id": q153, "selected_option": "D"}, headers=ch)
    await client.post("/api/exam/submit", headers=ch)

    report = await report_service.build_report(db, sitting, exam.name)
    assert [q["code"] for q in report["questions"]] == ["item_Q-2", "item_Q-81", "item_Q-153"]
    assert [q["correct_option"] for q in report["questions"]] == ["B", "C", "A"]
    assert [q["index"] for q in report["questions"]] == [1, 2, 3]
    # Đáp án thí sinh bám theo cột đã sắp: chọn D ở câu mã 153 → cột CUỐI.
    row = next(r for r in report["rows"] if r["cccd"] == cand.cccd)
    assert row["answers"] == ["", "", "D"]


async def test_report_keeps_order_when_codes_not_numeric(factory, db):
    """Đề không có mã (hoặc mã không kết thúc bằng số) → giữ nguyên thứ tự gốc."""
    admin, _ = await factory.admin()
    exam, sitting, _ = await factory.active_exam(
        [{"text": "Q1", "correct": "A"}, {"text": "Q2", "correct": "B"}],
        owner_id=admin.id,
    )
    report = await report_service.build_report(db, sitting, exam.name)
    assert [q["correct_option"] for q in report["questions"]] == ["A", "B"]
