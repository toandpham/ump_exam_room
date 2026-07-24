"""Giữ SỐ 0 ĐẦU của CCCD/hộ chiếu qua Excel (sự cố thực địa 15-07).

Excel tự nuốt số 0 đầu khi ô ở dạng Số. Người nhập danh sách chống lại bằng cách gõ
dấu nháy đơn ép ô thành chữ: ``'0123456789``. Phần mềm phải:
  - lúc NHẬP: hiểu dấu nháy đó là ký hiệu ép-chữ → bỏ đi, giữ nguyên số 0 đầu;
  - lúc XUẤT: ép ô về dạng Chữ + bật quotePrefix (chính là dấu ') → Excel không nuốt
    số 0, kể cả khi người dùng copy sang chỗ khác / lưu lại.
"""
import io

from openpyxl import Workbook, load_workbook

from app.services import excel_service, report_service


# ─── NHẬP: đọc hiểu dấu ' ────────────────────────────────────────────────────

def test_parse_cccd_strips_quote_prefix_keeps_leading_zero():
    # dấu nháy lọt thẳng vào giá trị (dán từ nơi khác / qua CSV)
    assert excel_service._parse_cccd("'012345678901") == "012345678901"
    # dấu nháy cong (Word/Zalo hay tự đổi)
    assert excel_service._parse_cccd("’012345678901") == "012345678901"
    # hộ chiếu cũng vậy
    assert excel_service._parse_cccd("'C1234567") == "C1234567"


def test_parse_cccd_keeps_plain_text_untouched():
    assert excel_service._parse_cccd("012345678901") == "012345678901"
    assert excel_service._parse_cccd("  079200000099  ") == "079200000099"


def test_parse_cccd_repads_when_excel_ate_the_zero():
    """Người nhập QUÊN dấu nháy → Excel biến thành SỐ, mất số 0 → bù lại cho đủ 12."""
    assert excel_service._parse_cccd(12345678901) == "012345678901"
    assert excel_service._parse_cccd(12345678901.0) == "012345678901"


def test_validate_row_accepts_quoted_cccd():
    """Cả dòng có dấu nháy vẫn qua validate + nhận đúng loại giấy tờ."""
    row = ("'012345678901", "Nguyễn Văn A", "1995-05-20", "Đơn vị A",
           2017, "Y đa khoa", "Đối tượng 1", 1, "Phòng 1")
    data, errors = excel_service.validate_row(row)
    assert errors == [], errors
    assert data["cccd"] == "012345678901"
    assert data["id_type"] == "cccd"


def test_validate_row_accepts_quoted_passport():
    row = ("'C1234567", "John Smith", "1990-08-15", "Đơn vị B",
           2015, "Y đa khoa", "Đối tượng 2", 1, None)
    data, errors = excel_service.validate_row(row)
    assert errors == [], errors
    assert data["cccd"] == "C1234567"
    assert data["id_type"] == "passport"


def test_import_roundtrip_from_real_xlsx_with_quote_prefix():
    """Ghi 1 file .xlsx y như người dùng gõ 'CCCD rồi đọc lại qua parse_rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(excel_service.HEADERS)
    ws.append(["'012345678901", "Nguyễn Văn A", "1995-05-20", "Đơn vị A",
               2017, "Y đa khoa", "Đối tượng 1", 1, "Phòng 1"])
    bio = io.BytesIO()
    wb.save(bio)

    rows = excel_service.parse_rows(bio.getvalue())
    assert len(rows) == 1
    data, errors = excel_service.validate_row(rows[0][1])
    assert errors == []
    assert data["cccd"] == "012345678901"   # số 0 đầu còn nguyên, không còn dấu nháy


# ─── XUẤT: ép Chữ + dấu ' ────────────────────────────────────────────────────

def _load(blob: bytes):
    return load_workbook(io.BytesIO(blob))


def test_template_forces_text_on_id_column():
    ws = _load(excel_service.build_template()).active
    for r in (2, 3):   # 2 dòng mẫu
        assert ws.cell(row=r, column=1).number_format == "@"
        assert ws.cell(row=r, column=1).quotePrefix is True


def test_export_candidates_forces_text_on_id_column():
    blob = excel_service.export_candidates([
        {"cccd": "012345678901", "full_name": "Nguyễn Văn A", "birth_date": "1995-05-20",
         "unit": "Đơn vị A", "category": "ĐT1", "attempt_number": 1},
    ])
    ws = _load(blob).active
    c = ws.cell(row=2, column=1)
    assert c.value == "012345678901"      # đọc lại vẫn đủ số 0
    assert c.number_format == "@"
    assert c.quotePrefix is True


def test_report_forces_text_on_id_column_both_sheets():
    """Cả sheet Kết quả lẫn Đáp án đều phải giữ số 0 đầu."""
    data = {
        "meta": {"exam_name": "Kỳ thi X", "sitting_name": "Buổi 1",
                 "exam_date": "2026-07-15", "question_count": 2},
        "rows": [{
            "stt": 1, "cccd": "012345678901", "ho_dem": "Nguyễn Văn", "ten": "A",
            "room_name": "Phòng 1", "birth_date": "1995-05-20", "status": "submitted",
            "score": 10.0, "total_correct": 2, "answers": ["A", "B"],
        }],
        "questions": [
            {"index": 1, "code": "Q-1", "correct_option": "A"},
            {"index": 2, "code": "Q-2", "correct_option": "B"},
        ],
    }
    wb = _load(report_service.export_excel(data))

    kq = wb["Kết quả"]
    c1 = kq.cell(row=6, column=2)          # dòng dữ liệu đầu, cột CCCD/Hộ chiếu
    assert c1.value == "012345678901"
    assert c1.number_format == "@" and c1.quotePrefix is True

    da = wb["Đáp án"]
    c2 = da.cell(row=8, column=2)          # dòng dữ liệu đầu, cột CCCD/Hộ chiếu (sheet Đáp án)
    assert c2.value == "012345678901"
    assert c2.number_format == "@" and c2.quotePrefix is True
